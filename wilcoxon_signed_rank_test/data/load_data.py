from openpyxl import load_workbook


def load_data(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.get_sheet_by_name('Data')

    data = []
    for row in sheet.iter_rows(min_row=2,
                               min_col=1,
                               max_col=7,
                               values_only=True):
        data.append({
            "customer": row[0],
            48806: row[1],
            47106: row[2],
            47287: row[3],
            48020: row[4],
            48863: row[5],
            50546: row[6]
        })
    return data

