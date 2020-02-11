from openpyxl import load_workbook


def load_data(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.get_sheet_by_name('Data')

    data = []

    names = list(sheet.iter_rows(min_row=1, max_row=1, min_col=1, values_only=True))[0]

    # todo for vs list()
    for row in sheet.iter_rows(min_row=2,
                               min_col=1,
                               values_only=True):
        data.append(row)
    return names, data

